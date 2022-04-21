import pandas as pd
import ast
from openpyxl import load_workbook
from csv import writer
import fileinput
import csv

"""
--------------------------------------------
class Excel_Data
--------------------------------------------
Here you will find everything related to the manage information in the csv and xls files.
"""


class Excel_Data:

    def __init__(self):
        self._tab_menu_atrib_values = []  # value for the tab defined in the excel
        self._tab_menu_atrib_names = []  # names for the tab defined in the excel
        self._tab_menu_atrib_layout = []  # type of object(combo,radio,check) for the tab data defined in the excel
        self._tab_menu_index = []  # amount the objects in the tab
        self._phrases_data_names = {}  # data save in the excel
        self._phrases_data = {}  # data save in the excel
        self._csv_url = None  # csv url
        self._sheet_names = None  # sheet names
        self._excel_url = None  # excel url

    # --------------------------------------------
    # getters and setters for this class
    # --------------------------------------------
    @property
    def excel_url(self):
        return self._excel_url

    @excel_url.setter
    def excel_url(self, excel_url_v):
        self._excel_url = excel_url_v

    @property
    def csv_url(self):
        return self._csv_url

    @csv_url.setter
    def csv_url(self, csv_url_v):
        self._csv_url = csv_url_v

    @property
    def sheet_names(self):
        return self._sheet_names

    @sheet_names.setter
    def sheet_names(self, sheet_names_v):
        self._sheet_names = sheet_names_v

    @property
    def tab_menu_index(self):
        return self._tab_menu_index

    @tab_menu_index.setter
    def tab_menu_index(self, tab_menu_index_v):
        self._tab_menu_index = tab_menu_index_v

    @property
    def phrases_data(self):
        return self._phrases_data

    @phrases_data.setter
    def phrases_data(self, phrases_data_v):
        self._phrases_data = phrases_data_v

    @property
    def phrases_data_names(self):
        return self._phrases_data_names

    @phrases_data_names.setter
    def phrases_data_names(self, phrases_data_names_v):
        self._phrases_data_names = phrases_data_names_v

    @property
    def tab_menu_atrib_values(self):
        if self.excel_url is None:
            return print("NO hay ruta del excel")
        else:
            xls = pd.ExcelFile(self.excel_url)
            self.sheet_names = xls.sheet_names
            for n in range(0, len(self.sheet_names)):
                df = pd.read_excel(xls, self.sheet_names[n])
                df_index = df.index.stop
                self.tab_menu_index.append(df.index.stop)
                data = pd.DataFrame(df, columns=['Value'])
                for n in range(0, df_index):
                    data_clean = self.data_cleanner(str(data.values[n]))
                    self._tab_menu_atrib_values.append(data_clean)

        return self._tab_menu_atrib_values

    @property
    def tab_menu_atrib_names(self):
        if self.excel_url is None:
            return print("NO hay ruta del excel")
        else:
            xls = pd.ExcelFile(self.excel_url)
            for n in range(0, len(self.sheet_names)):
                df = pd.read_excel(xls, self.sheet_names[n])
                df_index = df.index.stop
                data = pd.DataFrame(df, columns=['Name'])
                for n in range(0, df_index):
                    data_clean = self.data_cleanner(str(data.values[n]))
                    self._tab_menu_atrib_names.append(data_clean)

        return self._tab_menu_atrib_names

    @property
    def tab_menu_atrib_layout(self):
        if self.excel_url is None:
            return print("NO hay ruta del excel")
        else:
            xls = pd.ExcelFile(self.excel_url)
            for n in range(0, len(self.sheet_names)):
                df = pd.read_excel(xls, self.sheet_names[n])
                df_index = df.index.stop
                data = pd.DataFrame(df, columns=['Layout'])
                for n in range(0, df_index):
                    data_clean = self.data_cleanner(str(data.values[n]))
                    self._tab_menu_atrib_layout.append(data_clean)
        return self._tab_menu_atrib_layout

    # read dictionary data and convert to tab information
    def read_layout_phrases(self, index):
        xls = pd.ExcelFile(self.excel_url)
        df = pd.read_excel(xls, self.sheet_names[len(self.sheet_names) - 1])
        phrase = pd.DataFrame(df, columns=['Phrase'])
        system = pd.DataFrame(df, columns=['System'])
        print(self.to_dic_from_string(str(system.values[index])[2:-2:]))
        return (str(phrase.values[index])[2:-2:], self.to_dic_from_string(str(system.values[index])[2:-2:]))

    # write first time for that ID
    def write_phrase_data(self, tab_menu_attrib):
        # load excel file
        xls = pd.ExcelFile(self.excel_url)
        df = pd.read_excel(xls, self.sheet_names[len(self.sheet_names) - 1])
        df_index = df.index.stop
        wb = load_workbook(filename=self.excel_url)
        ws = wb.get_sheet_by_name(self.sheet_names[len(self.sheet_names) - 1])
        # write for first time data
        ws.cell(df_index + 2, 1).value = self.phrases_data[0]
        ws.cell(df_index + 2, 3).value = str(self.phrases_data_names)
        wb.save(self.excel_url)
        self.write_csv_data(tab_menu_attrib)

    # write csv first time for that ID
    def write_csv_data(self, tab_menu_attrib):
        i = 0
        temp = 0
        for n in range(0, len(tab_menu_attrib.sheet_names) - 1):
            ranges = tab_menu_attrib.tab_menu_index[i]
            for n2 in range(0, ranges):
                tab_menu_attrib_temp_titles = str(tab_menu_attrib.tab_menu_atrib_names[temp])
                list_data = [self.phrases_data[0], tab_menu_attrib.sheet_names[n], tab_menu_attrib_temp_titles[2:-2:],
                             self.phrases_data[temp + 1]]
                with open(self.csv_url, 'a', newline='') as f_object:
                    # Pass the CSV  file object to the writer() function
                    writer_object = writer(f_object)
                    # Result - a writer object
                    # Pass the data in the list as an argument into the writerow() function
                    writer_object.writerow(list_data)
                    # Close the file object
                    f_object.close()
                temp = temp + 1

            i = i + 1

    # modify data for that ID
    def modify_phrase_data(self, index, tab_menu_attrib):
        # load file
        wb = load_workbook(filename=self.excel_url)
        ws = wb.get_sheet_by_name(self.sheet_names[len(self.sheet_names) - 1])
        # modify the cells
        ws.cell(index + 2, 1).value = self.phrases_data[0]
        ws.cell(index + 2, 3).value = str(self.phrases_data_names)
        wb.save(self.excel_url)
        self.modify_csv_data(tab_menu_attrib)

    # modify csv data for that ID
    def modify_csv_data(self, tab_menu_attrib):
        temp = 0
        # modify the csv data
        with fileinput.input(files=self.csv_url, inplace=True, mode='r') as f:
            reader = csv.DictReader(f)
            print(",".join(reader.fieldnames))  # print back the headers
            ranges = tab_menu_attrib.tab_menu_index[0]
            n = 0
            flag = 0
            for row in reader:
                tab_menu_attrib_temp_titles = str(tab_menu_attrib.tab_menu_atrib_names[temp])
                # modify the row with the parameter ID equal to the getter
                if row["ID"] == self.phrases_data[0]:
                    row["Type"] = tab_menu_attrib.sheet_names[flag]
                    row['Name'] = tab_menu_attrib_temp_titles[2:-2:]
                    row['Value'] = self.phrases_data[temp + 1]
                    if n == ranges:
                        n = 0
                        flag = flag + 1
                    else:
                        n = n + 1
                    temp = temp + 1

                print(",".join([row["ID"], row["Type"], row['Name'], row['Value']]))

    # search if there are parameters saved with the searched id
    def search_id(self, id_person):

        xls = pd.ExcelFile(self.excel_url)
        df = pd.read_excel(xls, self.sheet_names[len(self.sheet_names) - 1])
        df_index = df.index.stop
        data = pd.DataFrame(df, columns=['ID'])
        for n in range(0, df_index):
            data_clean = str(data.values[n])
            if data_clean[2:-2:] == id_person:
                print('entro')
                return (True, n)

        return (False, None)

    # convert string to dictionary
    def to_dic_from_string(self, string):
        for i in string.splitlines():
            dic = ast.literal_eval(i)
        return dic

    def data_cleanner(self, data):
        data_clean = data[2:-2:].split(',')
        return data_clean
